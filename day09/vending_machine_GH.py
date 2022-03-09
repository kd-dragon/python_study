import time
import tqdm as tq
import openpyxl

wb = openpyxl.load_workbook('drink_list_v02.xlsx')
ws = wb['Sheet1']

cell_range = ws.cell(row=1, column=5).value  # 현재 품목의 수 11
menu_cell = ws['A1':'A{0}'.format(cell_range)]  # 품목의 끝행까지 호출
price_cell = ws['B1':'B{0}'.format(cell_range)]  # 가격의 끝행까지 호출

menu_list = []
price = []
menu_range = 0
money_box = 0

print(type(menu_cell))

for row in menu_cell:
    for cell in row:
        menu_list.append(cell.value)

for row in price_cell:
    for cell in row:
        price.append(cell.value)

menu_range = cell_range


def delay_time(num):
    count = num
    # count 부터 0까지 -1씩 증감하는 for 문
    for x in range(count, 0, -1):
        time.sleep(1)
    print("\n")


def percent_gauge():
    total = 0
    for x in tq.tqdm(range(100), ncols=100, desc="프로그램 시작 ="):
        time.sleep(0.001)
        total += x
    print(total)


def add_menu():  ###### 음료추가
    global cell_range
    global menu_range

    ws['E1'] = cell_range + 1

    a = input('추가할 음료는?:')
    menu_list.append(a)
    ws['A{0}'.format(menu_range + 1)] = a

    time.sleep(1)

    b = int(input('음료의 가격은?(백원단위로 입력):'))
    price.append(b)
    ws['B{0}'.format(menu_range + 1)] = b

    time.sleep(1)
    wb.save('drink_list_v02.xlsx')
    time.sleep(2)

    menu_range += 1


def menu_chart():
    print("\n\n")
    print("음료목록=================================================")
    print("\n")
    for i in range(0, menu_range):
        print("(", i + 1, ")", menu_list[i], ":", price[i], "원")


def del_menu():  ##### 음료삭제

    global menu_range

    a = input('삭제할 음료는?:')

    i = menu_list.index(a)

    del menu_list[i]
    del price[i]
    menu_range -= 1


def input_coin():  ###### 자판기 이용
    global money_box

    while True:

        try:
            a = int(input('동전을 넣으세요(500원이상):'))

            if a <= 499:

                print("\n\n")
                print("500원 이상만 넣어주세요!")
                print("\n")

            elif a % 100 > 0:

                print("\n\n")
                print("백원 단위로 넣어주세요~")
                print("\n")

            else:

                break

        except ValueError:
            print("\n\n")
            print('잘못된 입력입니다. (문자사용).')
            delay_time(1)
            a = 0
            break

    money_box += a

    print("\n\n")
    menu_chart()
    print("\n")
    print("(", "현재 금액:", money_box, ")")
    print("\n")


def buy_menu():
    global money_box

    try:
        while True:

            b = int(input('음료를 선택하세요(순서대로 1, 2, 3... 입력, 거스름돈 받기: 0 입력):'))
            print()

            if b >= 0:  # 0 값을 입력했을 때 음수값이 되는것을 피하기 위한 줄
                b -= 1

            else:  # 음수값이 아니라면 그대로 내보낸다.
                b = b

            if b >= 0 and len(menu_list) > b:

                if money_box >= 500 and money_box >= price[b]:
                    money_box -= price[b]

                    delay_time(1)
                    print("*", menu_list[b], "*", "가 나왔습니다.")
                    delay_time(1)
                    print("가격은", price[b], "원 입니다.")
                    delay_time(1)
                    menu_chart()
                    print("\n")
                    print("(", "현재 금액:", money_box, ")")



                elif b > 0:

                    print("\n\n")
                    menu_chart()
                    print("\n\n")
                    print("잔돈이 부족합니다.", "(", "현재 금액:", money_box, ")")
                    print("\n")


            elif b == -1:

                print("\n\n")
                print('거스름돈은', money_box, '원 입니다.')
                print("\n")

                money_box = 0

                break


            elif b < -1:
                print("\n")
                print("(", "현재 금액:", money_box, ")")
                print("\n\n")
                print('잘못된 입력입니다. (음수값).')
                delay_time(1)
                print('거스름돈은', money_box, '원 입니다.')
                print("\n")

                money_box = 0
                break

            else:

                print("\n\n")
                menu_chart()
                print("\n")
                print("(", "현재 금액:", money_box, ")")
                print("\n\n")
                print('메뉴가 없습니다.')
                print("\n")
    except ValueError:
        print("\n\n")
        print('잘못된 입력입니다. (문자사용).')
        delay_time(1)
        print('거스름돈은', money_box, '원 입니다.')
        print("\n")

        money_box = 0


# 게이지 바


percent_gauge()

while True:

    # 프로그램 실행

    print("======================기능==============================")
    print()
    print("1: 음료추가 2: 음료삭제 3: 자판기이용 4: 종료")
    print()
    print("========================================================")

    inputdata = input('기능선택 : ')

    if inputdata == '1':
        menu_chart()
        add_menu()
        menu_chart()

    elif inputdata == '2':
        menu_chart()
        del_menu()
        menu_chart()

    elif inputdata == '3':
        menu_chart()
        input_coin()
        buy_menu()
        delay_time(2)

    elif inputdata == '4':
        print()
        print('종료')
        print()
        break


    else:
        print("\n\n")
        print("없는 옵션인'대용'?")


