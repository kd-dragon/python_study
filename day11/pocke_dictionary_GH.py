import openpyxl

wb = openpyxl.load_workbook('pockemon.xlsx')
ws = wb['Sheet1']

xl_maxRow = ws.max_row
xl_num = ws['A1':'A{0}'.format(xl_maxRow)]
xl_name = ws['B1':'B{0}'.format(xl_maxRow)]
xl_deal = ws['D1':'D{0}'.format(xl_maxRow)]

poke_num = []
poke_name = []
poke_deal = []

for i in xl_num:
    for cell in i:
        poke_num.append(cell.value)

for i in xl_name:
    for cell in i:
        poke_name.append(cell.value)

for i in xl_deal:
    for cell in i:
        poke_deal.append(cell.value)


def poke_chart():
    print("\n\n")
    print("포켓몬 차트=================================================")
    print("\n")
    for i in range(1, len(poke_num)):
        print("(", "{0}".format(i).rjust(3), ")", ":", "{0}".format(poke_name[i]).center(2), "--DMG :".rjust(5),
              poke_deal[i])

    print("\n\n")


def poke_serch():
    a = input("찾을 포켓몬의 '번호' 또는 '이름'을 입력해 주세요!!! [현재 {0}종 ]  ".format(len(poke_num) - 1))

    try:
        i = poke_name.index(a)


    except ValueError:

        i = poke_name.index(poke_name[int(a)])

    while True:

        if len(poke_num) - 1 >= i > 0:
            print("\n\n")
            print("찾으시는 포켓몬 정보입니다.")
            print("\n")
            print("(", "{0}".format(i).rjust(3), ")", ":", "{0}".format(poke_name[i]).center(2), "--DMG :".rjust(5),
                  poke_deal[i])
            print("\n\n")
            break


        else:
            print("\n\n")
            print("없는 포켓몬 입니다.")
            print("\n")
            break


def poke_fight():
    a = input("[1] 싸울 포켓몬의 '번호' 또는 '이름'을 입력해 주세요!!! [현재 {0}종 ]  ".format(len(poke_num) - 1))

    try:

        ai = poke_name.index(a)


    except ValueError:

        ai = poke_name.index(poke_name[int(a)])

    b = input("[2] 싸울 포켓몬의 '번호' 또는 '이름'을 입력해 주세요!!! [현재 {0}종 ]  ".format(len(poke_num) - 1))

    try:

        bi = poke_name.index(b)

    except ValueError:

        bi = poke_name.index(poke_name[int(b)])

    while True:

        if poke_deal[ai] > poke_deal[bi]:

            print("\n\n")
            print("{0}".format(poke_name[ai]).center(2), "이/가 더 강합니다.")
            print("\n")
            print("(", "{0}".format(ai).rjust(3), ")", ":", "{0}".format(poke_name[ai]).center(2), "--DMG :".rjust(5),
                  poke_deal[ai], "WIN!!!")
            print("(", "{0}".format(bi).rjust(3), ")", ":", "{0}".format(poke_name[bi]).center(2), "--DMG :".rjust(5),
                  poke_deal[bi], "LOSE..")
            print("\n\n")
            break

        elif poke_deal[ai] < poke_deal[bi]:

            print("\n\n")
            print("{0}".format(poke_name[bi]).center(2), "이/가 더 강합니다.")
            print("\n")
            print("(", "{0}".format(bi).rjust(3), ")", ":", "{0}".format(poke_name[bi]).center(2), "--DMG :".rjust(5),
                  poke_deal[bi], "WIN!!!")
            print("(", "{0}".format(ai).rjust(3), ")", ":", "{0}".format(poke_name[ai]).center(2), "--DMG :".rjust(5),
                  poke_deal[ai], "LOSE..")
            print("\n\n")
            break

        elif ai == 0 or bi == 0:
            print("\n\n")
            print("??????")
            print("\n")
            break



        else:
            print("\n\n")
            print("??????")
            print("\n")
            break


# 게이지 바


while True:

    # 프로그램 실행

    print("======================기능==============================")
    print()
    print("1: 포켓몬 확인 2: 포켓몬 찾기 3: 포켓몬 시합 4: 종료 ")
    print()
    print("========================================================")

    inputdata = input('기능선택 : ')

    if inputdata == '1':
        poke_chart()


    elif inputdata == '2':
        poke_chart()
        poke_serch()




    elif inputdata == '3':
        poke_chart()
        poke_fight()



    elif inputdata == '4':
        print()
        print('종료')
        print()
        break


    else:
        print("\n\n")
        print("없는 옵션인'대용'?")


