import openpyxl as xls

workbook = xls.load_workbook('pockemon.xlsx')
sheet = workbook['Sheet1']

row_range = sheet.max_row
cell_range = sheet.max_column

pocket_list = []


class PocketMon:

    def set_data(self, num, name, hpoint, attack, defence, speed):
        self.num = num
        self.name = name
        self.hpoint = hpoint
        self.attack = attack
        self.defence = defence
        self.speed = speed

    def print_data(self):
        print(self.num, end=" | ")
        print(self.name, end=" | ")
        print(self.hpoint, end=" | ")
        print(self.attack, end=" | ")
        print(self.defence, end=" | ")
        print(self.speed)

    def append_list(self):
        obj_list = []

        obj_list.append(self.num)
        obj_list.append(self.name)
        obj_list.append(self.hpoint)
        obj_list.append(self.attack)
        obj_list.append(self.defence)
        obj_list.append(self.speed)

        pocket_list.append(obj_list)

    def __init__(self):
        self.num = ""
        self.name = ""
        self.hpoint = 0
        self.attack = 0
        self.defence = 0
        self.speed = 0


def get_pockemon_by_row():
    x = 1
    for i in range(2, row_range + 1):
        pocket_mon = PocketMon()

        for j in range(1, 10):
            cell_obj = sheet.cell(row=i, column=j)

            if j == 1:
                pocket_mon.num = cell_obj.value

            elif j == 2:
                pocket_mon.name = cell_obj.value

            elif j == 3:
                pocket_mon.hpoint = cell_obj.value

            elif j == 4:
                pocket_mon.attack = cell_obj.value

            elif j == 5:
                pocket_mon.defence = cell_obj.value

            elif j == 6:
                pocket_mon.speed = cell_obj.value

        pocket_list.append(pocket_mon)


def print_list():
    for a in pocket_list:
        print(a.num, end=" | ")
        print(a.name, end=" | ")
        print(a.hpoint, end=" | ")
        print(a.attack, end=" | ")
        print(a.defence, end=" | ")
        print(a.speed)


get_pockemon_by_row()

print_list()


def search():
    keyword = ""

    try:
        keyword = input("찾을 포켓몬의 '번호' 또는 '이름'을 입력해 주세요.")

        keyword = int(keyword)

        for obj in pocket_list:
            if obj.num == keyword:
                obj.print_data()
                return

    except ValueError:

        for obj in pocket_list:
            if keyword in obj.name:
                obj.print_data()


search()



